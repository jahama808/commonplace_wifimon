"""MDU↔OLT map parsing and lookup.

The source spreadsheet's `SAG` column contains MDU names embedded in a
free-form string. Examples we have to handle:

  *MDU - HERITAGE HOUSE*                                 → "HERITAGE HOUSE"
  *MDU - SEASIDE TOWERS*                                 → "SEASIDE TOWERS"
  FTTPB; CAF2A;MDU - TERRACES AT MANELE BAY PHASE I-III  → "TERRACES AT …"
  MDU - 1506 PIIKOI                                      → "1506 PIIKOI"

A single MDU appears multiple times (different OLTs / 7x50 redundancy),
so we de-dupe on the natural key (mdu_name, equip_name, equip_name_1).
"""
from __future__ import annotations

import io
import re
from typing import Any

import openpyxl
from sqlalchemy import delete, select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.mdu_olt_map import MduOltMap

_MDU_PREFIX = re.compile(r"\bMDU\s*-\s*", re.IGNORECASE)
# Only require the columns we actually consume; SYSTEM_NAME is in the
# source file but unused.
_REQUIRED_HEADERS = {
    "SAG",
    "FDH_NAME",
    "EQUIP_NAME",
    "SERVING_OLT",
    "EQUIP_NAME_1",
    "EQUIP_MODEL",
}


def extract_mdu_name(sag: str | None) -> str | None:
    """Return the MDU name embedded in a SAG string, or None.

    Strips leading prefix junk like 'FTTPB; CAF2A;' (anything before the
    first 'MDU - ') and trailing/leading asterisks + whitespace.
    """
    if not isinstance(sag, str):
        return None
    m = _MDU_PREFIX.search(sag)
    if not m:
        return None
    name = sag[m.end():].strip(" \t*")
    # Some rows have a trailing asterisk that the strip already handled,
    # but defend against doubled-up cases like "*MDU - X**".
    return name or None


def _normalize_cell(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def parse_xlsx(file_bytes: bytes) -> list[dict[str, str | None]]:
    """Parse the uploaded .xlsx and return one record per OLT row that
    has a recognizable MDU name.

    Raises ValueError on schema problems (wrong sheet, missing columns)
    so the caller can surface a 400 to the operator with a useful message.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:  # noqa: BLE001 — translate any openpyxl error
        raise ValueError(f"could not read xlsx: {e}") from e

    # Prefer 'Export Worksheet' (what the vendor exports), fall back to
    # the first sheet so a renamed export still works.
    ws = wb["Export Worksheet"] if "Export Worksheet" in wb.sheetnames else wb.active
    if ws is None:
        raise ValueError("xlsx has no readable sheet")

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError("xlsx has no header row")

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    missing = _REQUIRED_HEADERS - set(headers)
    if missing:
        raise ValueError(
            f"xlsx is missing required column(s): {sorted(missing)} "
            f"(have: {[h for h in headers if h]})"
        )

    idx = {h: i for i, h in enumerate(headers)}

    out: list[dict[str, str | None]] = []
    seen: set[tuple[str, str | None, str | None]] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        sag = row[idx["SAG"]] if idx["SAG"] < len(row) else None
        name = extract_mdu_name(sag)
        if not name:
            continue
        rec = {
            "mdu_name": name,
            "fdh_name": _normalize_cell(row[idx["FDH_NAME"]]),
            "equip_name": _normalize_cell(row[idx["EQUIP_NAME"]]),
            "serving_olt": _normalize_cell(row[idx["SERVING_OLT"]]),
            "equip_name_1": _normalize_cell(row[idx["EQUIP_NAME_1"]]),
            "equip_model": _normalize_cell(row[idx["EQUIP_MODEL"]]),
        }
        # Dedupe within the file so we don't fail the upsert for repeated
        # rows in the source.
        key = (rec["mdu_name"], rec["equip_name"], rec["equip_name_1"])
        if key in seen:
            continue
        seen.add(key)
        out.append(rec)

    return out


async def replace_all(session: AsyncSession, records: list[dict[str, str | None]]) -> int:
    """Wipe + re-insert. The spreadsheet is the source of truth — partial
    updates make no sense and would leave stale rows when an MDU is
    decommissioned upstream."""
    await session.execute(delete(MduOltMap))
    if records:
        await session.execute(pg_insert(MduOltMap), records)
    await session.commit()
    return len(records)


async def lookup_by_property_name(
    session: AsyncSession, property_name: str
) -> MduOltMap | None:
    """Best-effort match: the property name (admin-entered) matches an
    `mdu_name` (case-insensitive). Returns the first row, since multiple
    OLT rows for the same MDU are common in the source — display picks
    one for the central-office line."""
    if not property_name:
        return None
    q = (
        select(MduOltMap)
        .where(MduOltMap.mdu_name.ilike(property_name))
        .order_by(MduOltMap.id)
        .limit(1)
    )
    return (await session.execute(q)).scalar_one_or_none()


async def distinct_mdu_names(session: AsyncSession) -> list[str]:
    """For the property-create autocomplete. Sorted, deduped."""
    q = select(MduOltMap.mdu_name).distinct().order_by(MduOltMap.mdu_name)
    return [n for (n,) in (await session.execute(q)).all() if n]
